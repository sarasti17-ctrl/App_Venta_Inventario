from openpyxl import load_workbook
import os

try:
    file_path = os.path.join('data', 'Agujetas.xlsx')
    wb = load_workbook(filename=file_path, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        for row in sheet.iter_rows(max_row=5, values_only=True):
            print(row)
except Exception as e:
    print(f"Error: {e}")
